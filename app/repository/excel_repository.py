from pathlib import Path
from typing import Optional

import pandas as pd
from fastapi import HTTPException
from openpyxl import load_workbook

from app.config import EXCEL_FILE, DATA_DIR


def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = df.columns.astype(str).str.strip()
    return df


def none_safe_records(df: pd.DataFrame):
    df = df.where(pd.notnull(df), None)
    return df.to_dict(orient="records")


def ensure_workbook_exists():
    if not EXCEL_FILE.exists():
        raise HTTPException(
            status_code=404,
            detail=f"Excel file not found. Put PMIS.xlsx in: {DATA_DIR}",
        )


def read_sheet(sheet_name: str) -> pd.DataFrame:
    ensure_workbook_exists()
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
        return clean_columns(df)
    except ValueError:
        raise HTTPException(status_code=404, detail=f"Sheet not found: {sheet_name}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Cannot read sheet {sheet_name}: {e}")


def sheet_records(sheet_name: str):
    return none_safe_records(read_sheet(sheet_name))


def save_uploaded_workbook(file_bytes: bytes):
    DATA_DIR.mkdir(exist_ok=True)
    EXCEL_FILE.write_bytes(file_bytes)
    return str(EXCEL_FILE)


def update_rack_position(rack_id: str, drawing_id: Optional[str], x: float, y: float):
    ensure_workbook_exists()
    wb = load_workbook(EXCEL_FILE)
    if "Rack_Position" not in wb.sheetnames:
        raise HTTPException(status_code=404, detail="Sheet not found: Rack_Position")

    ws = wb["Rack_Position"]
    headers = {str(cell.value).strip(): idx + 1 for idx, cell in enumerate(ws[1]) if cell.value}

    required = ["Rack_ID", "X", "Y"]
    missing = [c for c in required if c not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing columns in Rack_Position: {missing}")

    rack_col = headers["Rack_ID"]
    x_col = headers["X"]
    y_col = headers["Y"]
    drawing_col = headers.get("Drawing_ID")

    found_row = None
    for row_idx in range(2, ws.max_row + 1):
        if str(ws.cell(row_idx, rack_col).value).strip() == rack_id:
            found_row = row_idx
            break

    if found_row is None:
        found_row = ws.max_row + 1
        ws.cell(found_row, rack_col).value = rack_id

    if drawing_id and drawing_col:
        ws.cell(found_row, drawing_col).value = drawing_id

    ws.cell(found_row, x_col).value = x
    ws.cell(found_row, y_col).value = y

    wb.save(EXCEL_FILE)
    return {"Rack_ID": rack_id, "Drawing_ID": drawing_id, "X": x, "Y": y}
